import pandas as pd # type: ignore
import math
from datetime import date
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Border, PatternFill, Side, Alignment # type: ignore
from openpyxl.styles import Font # type: ignore


yes = input("Enter Yesterday Folder Name: ")
yesLocation = "//192.168.1.231/Planning Internal/Capacity planning/Capacity Report/2025/10. Oct/" + str(yes) + "/"

today = input("Enter Today's Folder Name: ")
todLocation = "//192.168.1.231/Planning Internal/Capacity planning/Capacity Report/2025/10. Oct/" + str(today) + "/"

cur_month = 'Oct'
plan_month = 'Nov'
plan_month_end = '30'
plan_next_month = 'Dec'
plan_next_month_end = '31'

yes_buyer = pd.read_csv(yesLocation + "Buyer wise monthly plan qty.csv")
tod_buyer = pd.read_csv(todLocation + "Buyer wise monthly plan qty.csv")

curr_month_plan_qt = int(input("Enter current month's plan qty: "))

result_buyers = []
result = pd.DataFrame()
for index, row in yes_buyer.iterrows():
    if row['Buyer'] != '-':
        result_buyers.append(row['Buyer'])
for index, row in tod_buyer.iterrows():
    if row['Buyer'] not in result_buyers and row['Buyer'] != '-':
        result_buyers.append(row['Buyer'])
result_buyers.append('-')
result['Buyer Name'] = result_buyers

tod_first_t = today + '-' + cur_month + ' (' + plan_month + ')'
yes_first_t = yes + '-' + cur_month + ' (' + plan_month + ')'
change_first_t = 'Change (' + plan_month + ')'
tod_second_t = today + '-' + cur_month + ' (' + plan_next_month + ')'
yes_second_t = yes + '-' + cur_month + ' (' + plan_next_month + ')'
change_second_t = 'Change (' + plan_next_month + ')'
result[tod_first_t] = 0
result[yes_first_t] = 0
result[change_first_t] = 0
result[tod_second_t] = 0
result[yes_second_t] = 0
result[change_second_t] = 0

yes_first = {}
yes_second = {}
for index, row in yes_buyer.iterrows():
    yes_first[row['Buyer']] = 0 if math.isnan(row[1]) else row[1]
    yes_second[row['Buyer']] = 0 if math.isnan(row[2]) else row[2]

tod_first = {}
tod_second = {}
for index, row in tod_buyer.iterrows():
    tod_first[row['Buyer']] = 0 if math.isnan(row[1]) else row[1]
    tod_second[row['Buyer']] = 0 if math.isnan(row[2]) else row[2]

change_first = {}
change_second = {}
for buyer in result_buyers:
    if buyer == '-':
        continue
    if buyer not in tod_first.keys():
        change_first[buyer] = (-1) * yes_first[buyer]
    elif buyer not in yes_first.keys():
        change_first[buyer] = tod_first[buyer]
    else:
        change_first[buyer] = tod_first[buyer] - yes_first[buyer]

    if buyer not in tod_second.keys():
        change_second[buyer] = (-1) * yes_second[buyer]
    elif buyer not in yes_second.keys():
        change_second[buyer] = tod_second[buyer]
    else:
        change_second[buyer] = tod_second[buyer] - yes_second[buyer]

for index, row in result.iterrows():
    buyer = row['Buyer Name']
    if row['Buyer Name'] == '-':
        result.loc[index, tod_first_t] = tod_first[buyer] if buyer in tod_first.keys() else 0
        result.loc[index, yes_first_t] = yes_first[buyer] if buyer in yes_first.keys() else 0
        result.loc[index, tod_second_t] = tod_second[buyer] if buyer in tod_second.keys() else 0
        result.loc[index, yes_second_t] = yes_second[buyer] if buyer in yes_second.keys() else 0
    
        result.loc[index, change_first_t] = sum(change_first.values())
        result.loc[index, change_second_t] = sum(change_second.values())
    else:    
        result.loc[index, tod_first_t] = tod_first[buyer] if buyer in tod_first.keys() else 0
        result.loc[index, yes_first_t] = yes_first[buyer] if buyer in yes_first.keys() else 0
        result.loc[index, change_first_t] = change_first[buyer] if buyer in change_first.keys() else 0
        result.loc[index, tod_second_t] = tod_second[buyer] if buyer in tod_second.keys() else 0
        result.loc[index, yes_second_t] = yes_second[buyer] if buyer in yes_second.keys() else 0
        result.loc[index, change_second_t] = change_second[buyer] if buyer in change_second.keys() else 0

yes_unit = pd.read_csv(yesLocation + "Monthly blank days.csv")
tod_unit = pd.read_csv(todLocation + "Monthly blank days.csv")

tod_unit_cols = tod_unit.columns
yes_unit_cols = yes_unit.columns

result_unit = pd.DataFrame()
result_unit['Factory'] = tod_unit[['Factory']]
result_unit['-'] = tod_unit[['Unnamed: 1']]
result_unit[tod_first_t] = tod_unit[tod_unit_cols[2]]
result_unit[yes_first_t] = yes_unit[yes_unit_cols[2]]
result_unit[change_first_t] = result_unit[tod_first_t] - result_unit[yes_first_t]
result_unit[tod_second_t] = tod_unit[tod_unit_cols[3]]
result_unit[yes_second_t] = yes_unit[yes_unit_cols[3]]
result_unit[change_second_t] = result_unit[tod_second_t] - result_unit[yes_second_t]

first_w_days = 26
second_w_days = 26

first_blank_days = first_w_days * 427
second_blank_days = second_w_days * 427

first_capacity_pctg = (result_unit.iloc[7, 2] / first_blank_days) * 100
first_capacity_pctg = str(round(first_capacity_pctg, 2)) + '%'
second_capacity_pctg = (result_unit.iloc[7, 5] / second_blank_days) * 100
second_capacity_pctg = str(round(second_capacity_pctg, 2)) + '%'

# Adding a new row with loc (if the index is non-numeric, use a unique label)
result_unit.loc[8] = [None] * len(result_unit.columns)  # Add an empty row
result_unit.iloc[8, 2] = first_capacity_pctg
result_unit.iloc[8, 3] = first_blank_days
result_unit.iloc[8, 4] = str(first_w_days) + ' Days'

result_unit.iloc[8, 5] = second_capacity_pctg
result_unit.iloc[8, 6] = second_blank_days
result_unit.iloc[8, 7] = str(second_w_days) + ' Days'

provision = pd.read_csv(todLocation + 'Provision.csv')

provision_col_names = provision.columns

result_provision = pd.DataFrame()
for i in range(len(provision_col_names)):
    result_provision[provision_col_names[i]] = None

i = 0
# comment this for loop on the plan day
for index, row in provision.iterrows():
    one = row[provision_col_names[2]]
    two = row[provision_col_names[3]]
    three = row[provision_col_names[4]]
    four = row[provision_col_names[5]]
    if one > 0 or two > 0 or three > 0 or four > 0:
        for j in range(6):
            result_provision.loc[i, provision_col_names[j]] = row[provision_col_names[j]]
        i += 1

result_provision = result_provision.drop('Unnamed: 1', axis=1)

weekly_blank = pd.read_csv(todLocation + 'Weekly blank days.csv')

weekly_blank_col_names = weekly_blank.columns

result_weekly_blank = pd.DataFrame()
for index, row in weekly_blank.iterrows():
    for i in range(10):
        if i == 1:
            result_weekly_blank['-'] = weekly_blank[[weekly_blank_col_names[i]]]
        else:
            result_weekly_blank[weekly_blank_col_names[i]] = weekly_blank[[weekly_blank_col_names[i]]]

unit_and_buyer_wise = pd.read_csv(todLocation + 'Unit wise Buyer wise Plan Qty.csv')


# comparison
yes_comparison = pd.read_csv(yesLocation + 'Unit wise Buyer wise Plan Qty.csv')
tod_comparison = pd.read_csv(todLocation + 'Unit wise Buyer wise Plan Qty.csv')

comparison = pd.DataFrame()
comparison['Units'] = None
comparison['Yesterday Qty. (' + plan_month + ')'] = None
comparison['Today Qty. (' + plan_month + ')'] = None
comparison['Yesterday Qty. (' + plan_next_month + ')'] = None
comparison['Today Qty. (' + plan_next_month + ')'] = None
cnt = 0
for index, row in yes_comparison.iterrows():
    if row['Factory+Buyer'] == '-' and cnt < 8:
        comparison.loc[cnt, 'Units'] = row['Pl. Board']
        comparison.loc[cnt, 'Yesterday Qty. (' + plan_month + ')'] = row[2]
        comparison.loc[cnt, 'Yesterday Qty. (' + plan_next_month + ')'] = row[3]
        cnt += 1
cnt = 0
for index, row in tod_comparison.iterrows():
    if row['Factory+Buyer'] == '-' and cnt < 8:
        comparison.loc[cnt, 'Today Qty. (' + plan_month + ')'] = row[2]
        comparison.loc[cnt, 'Today Qty. (' + plan_next_month + ')'] = row[3]
        cnt += 1

comparison['Change (' + plan_month + ')'] = comparison['Today Qty. (' + plan_month + ')'] - comparison['Yesterday Qty. (' + plan_month + ')']
comparison['Change (' + plan_next_month + ')'] = comparison['Today Qty. (' + plan_next_month + ')'] - comparison['Yesterday Qty. (' + plan_next_month + ')']


today_date = date.today()
today_date = today_date.strftime("%d-%b-%y") # Example: 10-Mar-25
outputFile = str(today_date) + '.xlsx'
with pd.ExcelWriter(outputFile) as writer:
    result_unit.to_excel(writer, sheet_name='Unit Wise', index=False)
    result_weekly_blank.to_excel(writer, sheet_name="Weekly Blank Days", index=False)
    result.to_excel(writer, sheet_name='Buyer Wise', index=False)
    result_provision.to_excel(writer, sheet_name="Provision", index=False)
    unit_and_buyer_wise.to_excel(writer, sheet_name='Unit and Buyer Wise', index=False),
    comparison.to_excel(writer, sheet_name='Comparison', index=False)

wb = load_workbook(outputFile)
ws_unit = wb['Unit Wise']
ws_weekly_blank = wb['Weekly Blank Days']
ws_buyer = wb['Buyer Wise']
ws_provision = wb['Provision']
ws_unit_and_buyer = wb['Unit and Buyer Wise']
ws_comparison = wb['Comparison']

# define font to Arial
times_roman = Font(name='Arial')

# change the font
for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = times_roman


# Define the border style
thin_border = Border(left=Side(style='thin', color='FFa4a4a4'), 
                     right=Side(style='thin', color='FFa4a4a4'), 
                     top=Side(style='thin', color='FFa4a4a4'), 
                     bottom=Side(style='thin', color='FFa4a4a4'))

# Define the bold font style
bold_font = Font(name='Arial', bold=True,)

# Define the font with a specific color (e.g., red)
red_font = Font(name='Arial', color='FF0000', bold=True)

# comma
comma = '#,##0'

# title color
fill_color = PatternFill(start_color='dce1e0', end_color='dce1e0', fill_type='solid')

def is_number(cell):
    return isinstance(cell.value, (int, float))

# Apply the border to all cells in the DataFrame
for row in ws_unit.iter_rows(min_row=1, max_row=ws_unit.max_row, min_col=1, max_col=ws_unit.max_column):
    for cell in row:
        cell.border = thin_border
        if is_number(cell) == True:
            cell.number_format = comma

for row in ws_weekly_blank.iter_rows(min_row=1, max_row=ws_weekly_blank.max_row, min_col=1, max_col=ws_weekly_blank.max_column):
    for cell in row:
        cell.border = thin_border
        if is_number(cell) == True:
            cell.number_format = comma

for row in ws_buyer.iter_rows(min_row=1, max_row=ws_buyer.max_row, min_col=1, max_col=ws_buyer.max_column):
    for cell in row:
        cell.border = thin_border
        if is_number(cell) == True:
            cell.number_format = comma
        
for row in ws_provision.iter_rows(min_row=1, max_row=ws_provision.max_row, min_col=1, max_col=ws_provision.max_column):
    for cell in row:
        cell.border = thin_border
        if is_number(cell) == True:
            cell.number_format = comma

for row in ws_unit_and_buyer.iter_rows(min_row=1, max_row=ws_unit_and_buyer.max_row, min_col=1, max_col=ws_unit_and_buyer.max_column):
    for cell in row:
        cell.border = thin_border
        if is_number(cell) == True:
            cell.number_format = comma

count = 0
for row in ws_unit.iter_rows():
    for cell in row:
        if count == 0 or count == 8 or count == 9:
            cell.font = bold_font
            cell.fill = fill_color
        if is_number(cell) == True:
            cell.number_format = comma
    count += 1

count = 0
for row in ws_comparison.iter_rows():
    for cell in row:
        cell.border = thin_border
        if count == 0:
            cell.font = bold_font
            cell.fill = fill_color
        val = cell.value
        if isinstance(val, (int, float)):
            if val < 0:
                cell.font = red_font
            cell.number_format = comma
    count += 1

ws_unit.column_dimensions['B'].width = 20
ws_unit.column_dimensions['C'].width = 13
ws_unit.column_dimensions['D'].width = 13
ws_unit.column_dimensions['E'].width = 13
ws_unit.column_dimensions['F'].width = 13
ws_unit.column_dimensions['G'].width = 13
ws_unit.column_dimensions['H'].width = 13

ws_comparison.column_dimensions['B'].width = 15
ws_comparison.column_dimensions['C'].width = 13
ws_comparison.column_dimensions['D'].width = 13

count = 0
for row in ws_weekly_blank.iter_rows():
    if count == 0 or count == 8:
        for cell in row:
            cell.font = bold_font
            cell.fill = fill_color
    count += 1

# red mark to the negative values in weekly blank days
for row in ws_weekly_blank.iter_rows():
    for cell in row:
        val = cell.value
        if isinstance(val, (int, float)):  # Check if the value is a number
            if val < 0:
                cell.font = red_font

count = 0
for row in ws_buyer.iter_rows():
    if count == 0 or ws_buyer.max_row == count + 1:
        for cell in row:
            cell.font = bold_font
            cell.fill = fill_color
    count += 1

# red mark to the negative values in buyer wise
for row in ws_buyer.iter_rows():
    for cell in row:
        val = cell.value
        if isinstance(val, (int, float)):  # Check if the value is a number
            if val < 0:
                cell.font = red_font

count = 0
for row in ws_provision.iter_rows():
    if count == 0 or ws_provision.max_row == count + 1:
        for cell in row:
            cell.font = bold_font
            cell.fill = fill_color
    count += 1

count = 0
for row in ws_unit_and_buyer.iter_rows(min_row=1, max_row=ws_unit_and_buyer.max_row, min_col=1, max_col=ws_unit_and_buyer.max_column):
    if row[1].value == '-' or count == 0:
        for cell in row:
            cell.font = bold_font
            cell.fill = fill_color
    count += 1

# red marked text to the negative values
for row in ws_unit.iter_rows():
    for cell in row:
        val = cell.value
        if isinstance(val, (int, float)):
            if val < 0:
                cell.font = red_font


ws_weekly_blank_range = ws_weekly_blank['A1:J9']
ws_unit['A15'] = 'Weekly Blank Days (Factory wise)'
ws_unit['A15'].font = Font(bold=True, name='Arial', size=14)
start_row = 16
start_col = 1
for row_idx, row in enumerate(ws_weekly_blank_range, start=start_row):
    for col_idx, cell in enumerate(row, start=start_col):
        destination_cell = ws_unit.cell(row=row_idx, column=col_idx, value=cell.value)
        # Copy the cell font
        if cell.has_style:
            destination_cell.font = Font(name=cell.font.name, 
                                         size=cell.font.size, 
                                         bold=cell.font.bold, 
                                         italic=cell.font.italic, 
                                         vertAlign=cell.font.vertAlign, 
                                         underline=cell.font.underline, 
                                         strike=cell.font.strike, 
                                         color=cell.font.color)
        # Copy the fill (cell background color)
        if cell.fill is not None:
            destination_cell.fill = PatternFill(fill_type=cell.fill.fill_type, 
                                                start_color=cell.fill.start_color, 
                                                end_color=cell.fill.end_color)
        # Copy the border
        if cell.border is not None:
            destination_cell.border = Border(left=cell.border.left,
                                                right=cell.border.right,
                                                top=cell.border.top,
                                                bottom=cell.border.bottom)
        # Copy number format
        if cell.number_format:
            destination_cell.number_format = cell.number_format

ws_buyer_range = ws_buyer['A1:G26']
ws_unit['B27'] = 'Buyer wise Monthly Plan qty.'
ws_unit['B27'].font = Font(bold=True, name='Arial', size=14)
start_row = 28
start_col = 2
for row_idx, row in enumerate(ws_buyer_range, start=start_row):
    for col_idx, cell in enumerate(row, start=start_col):
        destination_cell = ws_unit.cell(row=row_idx, column=col_idx, value=cell.value)
        # Copy the cell font
        if cell.has_style:
            destination_cell.font = Font(name=cell.font.name, 
                                         size=cell.font.size, 
                                         bold=cell.font.bold, 
                                         italic=cell.font.italic, 
                                         vertAlign=cell.font.vertAlign, 
                                         underline=cell.font.underline, 
                                         strike=cell.font.strike, 
                                         color=cell.font.color)
        # Copy the fill (cell background color)
        if cell.fill is not None:
            destination_cell.fill = PatternFill(fill_type=cell.fill.fill_type, 
                                                start_color=cell.fill.start_color, 
                                                end_color=cell.fill.end_color)
        # Copy the border
        if cell.border is not None:
            destination_cell.border = Border(left=cell.border.left,
                                                right=cell.border.right,
                                                top=cell.border.top,
                                                bottom=cell.border.bottom)
        # Copy number format
        if cell.number_format:
            destination_cell.number_format = cell.number_format

ws_provision_range = ws_provision['A1:E8']
ws_unit['B54'] = 'Buyer wise Monthly Provision'
ws_unit['B54'].font = Font(bold=True, name='Arial', size=14)
start_row = 55
start_col = 2
for row_idx, row in enumerate(ws_provision_range, start=start_row):
    for col_idx, cell in enumerate(row, start=start_col):
        destination_cell = ws_unit.cell(row=row_idx, column=col_idx, value=cell.value)
        # Copy the cell font
        if cell.has_style:
            destination_cell.font = Font(name=cell.font.name, 
                                         size=cell.font.size, 
                                         bold=cell.font.bold, 
                                         italic=cell.font.italic, 
                                         vertAlign=cell.font.vertAlign, 
                                         underline=cell.font.underline, 
                                         strike=cell.font.strike, 
                                         color=cell.font.color)
        # Copy the fill (cell background color)
        if cell.fill is not None:
            destination_cell.fill = PatternFill(fill_type=cell.fill.fill_type, 
                                                start_color=cell.fill.start_color, 
                                                end_color=cell.fill.end_color)
        # Copy the border
        if cell.border is not None:
            destination_cell.border = Border(left=cell.border.left,
                                                right=cell.border.right,
                                                top=cell.border.top,
                                                bottom=cell.border.bottom)
        # Copy number format
        if cell.number_format:
            destination_cell.number_format = cell.number_format

ws_unit_and_buyer_range = ws_unit_and_buyer['A1:G70']
ws_unit['A65'] = 'Unit wise, Buyer wise Monthly Plan Qty.'
ws_unit['A65'].font = Font(bold=True, name='Arial', size=14)
start_row = 66
start_col = 1
for row_idx, row in enumerate(ws_unit_and_buyer_range, start=start_row):
    for col_idx, cell in enumerate(row, start=start_col):
        destination_cell = ws_unit.cell(row=row_idx, column=col_idx, value=cell.value)
        # Copy the cell font
        if cell.has_style:
            destination_cell.font = Font(name=cell.font.name, 
                                         size=cell.font.size, 
                                         bold=cell.font.bold, 
                                         italic=cell.font.italic, 
                                         vertAlign=cell.font.vertAlign, 
                                         underline=cell.font.underline, 
                                         strike=cell.font.strike, 
                                         color=cell.font.color)
        # Copy the fill (cell background color)
        if cell.fill is not None:
            destination_cell.fill = PatternFill(fill_type=cell.fill.fill_type, 
                                                start_color=cell.fill.start_color, 
                                                end_color=cell.fill.end_color)
        # Copy the border
        if cell.border is not None:
            destination_cell.border = Border(left=cell.border.left,
                                                right=cell.border.right,
                                                top=cell.border.top,
                                                bottom=cell.border.bottom)
        # Copy number format
        if cell.number_format:
            destination_cell.number_format = cell.number_format

curr_month_plan_qt_cell = ws_unit['A12']
curr_month_plan_qt_cell.value = cur_month + " Plan Quantity ="
curr_month_plan_qt_cell.font = Font(name='Arial', bold=True)
ws_unit.merge_cells('A12:B12')

curr_month_plan_qt_value_cell = ws_unit['C12']
curr_month_plan_qt_value_cell.value = curr_month_plan_qt
curr_month_plan_qt_value_cell.font = Font(name='Arial', bold=True)
curr_month_plan_qt_value_cell.number_format = '#,##0'

ws_unit.insert_rows(1)
ws_unit['A1'] = 'Monthly Blank Days'
ws_unit['A1'].font = Font(bold=True, size=14, name='Arial')
ws_unit.insert_rows(1)
ws_unit['A1'] = 'Capacity Report'
ws_unit['A1'].font = Font(bold=True, size=24, name='Playfair Display')
ws_unit.merge_cells('A1:I1')
ws_unit['A1'].alignment = Alignment(horizontal='center', vertical='center')

wb.save(outputFile)
outputFile2 = '//192.168.1.231/Planning Internal/Capacity planning/Capacity Report/2025/Reports/10. Oct/' + str(today_date) + '.xlsx'
wb.save(outputFile2)

print("\nSuccessfully done :)")